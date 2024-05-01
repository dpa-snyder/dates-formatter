import pandas as pd
import datetime
from openpyxl import load_workbook

def excel_serial_to_date(serial):
    serial_str = str(serial)
    # check if serial is a five digit number
    if pd.isna(serial) or not serial_str.isdigit() or len(serial_str) != 5:
        return serial  # Return the original data if conditions are not met
    excel_start_date = datetime.date(1899, 12, 31)
    
    serial_int = int(serial_str)
    if serial_int > 59:
        serial_int += 1  # Excel leap year bug, 1900 is not a leap year

    # Convert the date
    date_converted = excel_start_date + datetime.timedelta(days=serial_int - 1)
    
    return date_converted.strftime('%m/%d/%Y')  # Return the formatted date string

def add_converted_date_column(excel_file, sheet_name, serial_column):
    try:
        # Load workbook to check if the sheet exists
        book = load_workbook(excel_file)
        if sheet_name not in book.sheetnames:
            print(f"No sheet named '{sheet_name}' found.")
            return

        # Load data from specified sheet
        data = pd.read_excel(excel_file, sheet_name=sheet_name)

        # Convert the serial dates or copy ignored values as-is
        data['Converted Date'] = data[serial_column].apply(excel_serial_to_date)

        # Write back to Excel without the index on the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    except Exception as e:
        print("An error occurred:", e)

# Usage
file_path = 'D:\\GitHub\\leading-zeros-dates\\fixme.xlsx'
sheet = 'Sheet1'
serial_date_column = 'FullDate'

add_converted_date_column(file_path, sheet, serial_date_column)
